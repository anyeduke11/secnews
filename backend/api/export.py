"""Phase 4 /api/export router — 预生成 HTML + ETag 304。"""
from __future__ import annotations

import asyncio
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Header, Response

from backend.logging_config import logger
from backend.services.export_service import (
    get_cached_etag,
    get_or_build_html,
)
from backend.version import APP_VERSION as API_VERSION

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/xlsx")
async def export_hotspots_xlsx(category: str | None = None):
    """导出热点数据为 XLSX 表格 (P0-7: 修复前端导出 404)。

    前端 ExportSettings 原调用 /api/export/download (不存在) → 404;
    新增此端点, 按分类导出最近 1000 条热点。
    """
    from backend.domain.enums import Category
    from backend.repository.hotspot_repo import HotspotRepository

    repo = HotspotRepository()
    # query 接受 Category 枚举, 字符串需转换
    cat_enum = None
    if category:
        try:
            cat_enum = Category(category)
        except ValueError:
            cat_enum = None
    # query 返回 (items, cursor) 元组
    items, _ = await asyncio.to_thread(
        repo.query, category=cat_enum, limit=1000
    )

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "热点清单"

    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["分类", "标题", "来源", "发布时间", "原文链接"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    cat_cn = {
        "ai": "科技/AI",
        "security": "网络安全",
        "finance": "金融/投资",
        "startup": "独立开发/创业",
        "bid": "招标资讯",
        "github": "GitHub 项目",
        "tech": "科技",
        "ai_security": "AI 安全",
    }
    body_font = Font(name="Microsoft YaHei", size=10)
    link_font = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
    for idx, it in enumerate(items, start=2):
        cat = getattr(it, "category", "") or ""
        a = ws.cell(row=idx, column=1, value=cat_cn.get(cat, cat))
        a.font = body_font
        a.alignment = center
        b = ws.cell(row=idx, column=2, value=getattr(it, "title", ""))
        b.font = body_font
        b.alignment = left
        ws.cell(row=idx, column=3, value=getattr(it, "source", "")).font = body_font
        ws.cell(row=idx, column=4, value=str(getattr(it, "published_at", "") or "")).font = body_font
        url = str(getattr(it, "url", "") or "")
        c = ws.cell(row=idx, column=5, value=url)
        c.hyperlink = url
        c.font = link_font
        c.alignment = left

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 55
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    cat_tag = category or "all"
    filename = f"hotspots_{cat_tag}_{ts}.xlsx"
    return Response(
        content=payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("")
async def export(
    if_none_match: str | None = Header(default=None, alias="If-None-Match"),
):
    """返回预生成的静态 HTML；客户端传 If-None-Match 触发 304。

    Phase 9 修复：cache miss 时同步 DB query + 文件 IO 放 thread pool。
    """
    etag = get_cached_etag() or '"no-cache"'
    if if_none_match and if_none_match.strip() == etag:
        return Response(status_code=304, headers={"ETag": etag})
    html, fresh_etag = await asyncio.to_thread(get_or_build_html)
    return Response(
        content=html,
        media_type="text/html; charset=utf-8",
        headers={
            "ETag": fresh_etag,
            "Cache-Control": "public, max-age=1800",
        },
    )


@router.post("/rebuild")
async def export_rebuild():
    """强制重建（运维用）。

    Phase 9 修复：同步 DB query + 文件 IO 放 thread pool。
    """
    from backend.services.export_service import rebuild_export_cache

    try:
        etag = await asyncio.to_thread(rebuild_export_cache)
        return {"version": API_VERSION, "etag": etag, "status": "ok"}
    except Exception as e:
        logger.error(f"export rebuild failed: {e}")
        return Response(
            content=f"rebuild failed: {e}",
            status_code=500,
        )
