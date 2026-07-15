"""Phase 10 收藏 API 端点

路由清单
--------
- ``GET    /api/favorites``                 列表（支持按 category 筛选）
- ``POST   /api/favorites``                 添加收藏
- ``DELETE /api/favorites/{hotspot_id}``    取消收藏
- ``GET    /api/favorites/count``           按分类统计 + 总数
- ``GET    /api/favorites/export``          xlsx 导出（3 列: 信息类型/标题/原文链接）

xlsx 格式约定
-------------
- 3 列: ``信息类型`` / ``标题名称`` / ``原文链接``
- 表头加粗 + 居中 + 浅蓝底色
- 链接列设为可点击超链接
- 文件名: ``favorites_{YYYYMMDD}_{HHMMSS}.xlsx``
- 使用 openpyxl 写入,符合 .xlsx (Office Open XML) 标准

异步策略
--------
所有同步 DB / 文件 IO 操作通过 ``asyncio.to_thread`` 包装,避免阻塞
event loop。``openpyxl`` Workbook 写入为纯 Python IO,必须放线程池。
"""
from __future__ import annotations

import asyncio
import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from backend.domain.enums import Category
from backend.logging_config import logger
from backend.repository.favorite_repo import FavoriteItem, FavoriteRepository

router = APIRouter(prefix="/api/favorites", tags=["favorites"])


# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------
class AddFavoriteRequest(BaseModel):
    """添加收藏请求体。"""

    hotspot_id: str = Field(..., min_length=1, max_length=128, description="hotspot 唯一 ID")
    category: str = Field(..., description="6 大分类之一: ai/security/finance/startup/bid/github")
    title: str = Field(..., min_length=1, max_length=500, description="卡片标题")
    source: str = Field(..., min_length=1, max_length=200, description="信源名")
    url: str = Field(..., min_length=1, max_length=2000, description="原文链接")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _validate_category(category: str) -> str:
    """校验 category,失败抛 HTTPException(400)。返回标准化后的值。"""
    if not category or not category.strip():
        raise HTTPException(status_code=400, detail={"message": "category 不能为空"})
    try:
        return Category.from_str(category).value
    except Exception:
        valid = ", ".join(repr(c.value) for c in Category)
        raise HTTPException(
            status_code=400,
            detail={"message": f"未知 category {category!r}; 合法值: {valid}"},
        )


def _build_list_payload(category: Optional[str], limit: int) -> dict:
    """同步读取收藏列表（在 thread pool 内执行）。"""
    repo = FavoriteRepository()
    if category:
        cat_value = _validate_category(category)
    else:
        cat_value = None
    items: list[FavoriteItem] = repo.list(category=cat_value, limit=limit)
    return {
        "version": "1.2.0",
        "category": cat_value or "all",
        "total": repo.total(),
        "count": len(items),
        "items": [it.to_dict() for it in items],
    }


def _build_count_payload() -> dict:
    """同步统计收藏数量（在 thread pool 内执行）。"""
    repo = FavoriteRepository()
    return {
        "version": "1.2.0",
        "total": repo.total(),
        "by_category": repo.count_by_category(),
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@router.get("")
async def list_favorites(
    category: Optional[str] = Query(None, description="分类筛选; 不传=全部"),
    limit: int = Query(200, ge=1, le=1000, description="最多返回条数"),
):
    """按收藏时间倒序列出收藏项。支持按 category 筛选。

    - ``category=ai`` → 只返回科技/AI 类
    - ``category=bid`` → 只返回招标资讯
    - ``category=all`` 或不传 → 全部
    """
    return await asyncio.to_thread(_build_list_payload, category, limit)


@router.get("/count")
async def count_favorites():
    """按分类统计收藏数 + 总数。全 6 大分类都有 key(无收藏也返回 0)。"""
    return await asyncio.to_thread(_build_count_payload)


@router.post("")
async def add_favorite(req: AddFavoriteRequest):
    """添加收藏。已存在则返回 200 + created=False。"""
    cat_value = _validate_category(req.category)
    repo = FavoriteRepository()
    try:
        created, item = await asyncio.to_thread(
            repo.add,
            hotspot_id=req.hotspot_id.strip(),
            category=cat_value,
            title=req.title.strip(),
            source=req.source.strip(),
            url=req.url.strip(),
        )
    except Exception as e:
        logger.error(f"add favorite failed: {e}")
        raise HTTPException(status_code=500, detail={"message": f"添加失败: {e}"})

    # v1.4: sync to knowledge items (non-critical, must not break favorites flow)
    try:
        from backend.repository.knowledge_repo import knowledge_repo
        from backend.domain.knowledge_models import KnowledgeItem, now_iso
        from backend.services.data_cleaning import item_id_from_url
        import logging
        _klog = logging.getLogger("hotspot.favorites.webhook")
        fav_url = req.url.strip()
        fav_title = req.title.strip()
        if fav_url:
            item_id = item_id_from_url(fav_url)
            existing = knowledge_repo.get_item(item_id)
            if not existing:
                kitem = KnowledgeItem(
                    id=item_id,
                    title=fav_title or "Untitled",
                    source="secnews",
                    source_url=fav_url,
                    ingested_at=now_iso(),
                    updated_at=now_iso(),
                )
                knowledge_repo.upsert_item(kitem)
                _klog.info(f"favorite synced to knowledge: {item_id}")
    except Exception as e:
        _klog = logging.getLogger("hotspot.favorites.webhook")
        _klog.warning(f"favorite -> knowledge sync failed (non-critical): {e}")

    return {
        "status": "ok",
        "created": created,
        "item": item.to_dict(),
    }


@router.delete("/{hotspot_id}")
async def remove_favorite(hotspot_id: str):
    """取消收藏。返回删除行数 (0=本来就没收藏)。"""
    if not hotspot_id or not hotspot_id.strip():
        raise HTTPException(status_code=400, detail={"message": "hotspot_id 不能为空"})
    repo = FavoriteRepository()
    try:
        n = await asyncio.to_thread(repo.remove, hotspot_id.strip())
    except Exception as e:
        logger.error(f"remove favorite failed: {e}")
        raise HTTPException(status_code=500, detail={"message": f"取消失败: {e}"})
    return {
        "status": "ok",
        "hotspot_id": hotspot_id,
        "removed": n,
    }


# ---------------------------------------------------------------------------
# xlsx 导出
# ---------------------------------------------------------------------------
def _build_xlsx(category: Optional[str]) -> tuple[bytes, str, int]:
    """同步构建 xlsx 文件 bytes + 文件名 + 行数（在 thread pool 内执行）。

    文件结构:
    - Sheet 名: "收藏清单"
    - 表头 (3 列): 信息类型 / 标题名称 / 原文链接
    - 表头样式: 加粗 + 居中 + 浅蓝底色 (#cfe2f3)
    - 数据行: 按收藏时间倒序
    - 链接列: 设置 hyperlink 方便 Excel 直接点击
    - 列宽: A=15 / B=60 / C=50
    """
    repo = FavoriteRepository()
    if category:
        cat_value = _validate_category(category)
    else:
        cat_value = None
    items = repo.list(category=cat_value, limit=1000)

    wb = Workbook()
    ws = wb.active
    ws.title = "收藏清单"

    # 表头样式
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["信息类型", "标题名称", "原文链接"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 数据行
    body_font = Font(name="Microsoft YaHei", size=10)
    link_font = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
    for idx, it in enumerate(items, start=2):
        # A: 信息类型 (category 中文映射)
        cat_cn = {
            "ai": "科技/AI",
            "security": "网络安全",
            "finance": "金融/投资",
            "startup": "独立开发/创业",
            "bid": "招标资讯",
            "github": "GitHub 项目",
        }.get(it.category, it.category)
        a = ws.cell(row=idx, column=1, value=cat_cn)
        a.font = body_font
        a.alignment = center
        # B: 标题
        b = ws.cell(row=idx, column=2, value=it.title)
        b.font = body_font
        b.alignment = left
        # C: 链接 (hyperlink)
        c = ws.cell(row=idx, column=3, value=it.url)
        c.hyperlink = it.url
        c.font = link_font
        c.alignment = left

    # 列宽
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 55
    # 表头行高
    ws.row_dimensions[1].height = 26
    # 冻结表头
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    cat_tag = cat_value or "all"
    filename = f"favorites_{cat_tag}_{ts}.xlsx"
    return payload, filename, len(items)


@router.get("/export")
async def export_favorites(
    category: Optional[str] = Query(None, description="分类筛选; 不传=全部导出"),
):
    """批量导出收藏到 .xlsx 文件 (3 列: 信息类型/标题/原文链接)。

    使用 ``StreamingResponse`` 避免大文件驻留内存 → 一次性写完 buffer 再 stream
    """
    try:
        payload, filename, count = await asyncio.to_thread(_build_xlsx, category)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"xlsx export failed: {e}")
        raise HTTPException(status_code=500, detail={"message": f"导出失败: {e}"})

    logger.info(
        f"favorites xlsx export: count={count} category={category or 'all'} bytes={len(payload)}"
    )
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Favorite-Count": str(count),
    }
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


__all__ = ["router"]
